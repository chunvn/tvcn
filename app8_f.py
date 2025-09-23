import io
import re
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment

st.set_page_config(page_title="Merge TVCN with Khách hàng by ID", layout="centered")

st.title("Merge TVCN with Khách hàng by ID")
st.caption(
    "Upload .xlsx → merge by ID → keep only 'Đã báo cáo' → normalize SĐT as text → "
    "add 'Số lần gặp' → keep only required columns/order → sort with SĐT groups preserved → merge group cells."
)

uploaded = st.file_uploader("Upload your Excel (.xlsx)", type=["xlsx"])

# Defaults
tvcn_sheet_default = "TVCN"
kh_sheet_default   = "Khách hàng"

# ---------- Helpers ----------
def normalize_id(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s or None

def normalize_phone(val):
    """Treat SĐT as text: strip trailing .0, keep digits only (preserve leading zeros)."""
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    digits = "".join(re.findall(r"\d", s))
    return digits or None

def filter_reported_rows(df):
    """Keep only rows where Trạng thái == 'Đã báo cáo' (case-insensitive; trims spaces)."""
    if "Trạng thái" not in df.columns:
        raise ValueError("Column 'Trạng thái' not found in the TVCN sheet.")
    norm = df["Trạng thái"].astype(str).str.strip().str.lower()
    return df.loc[norm.eq("đã báo cáo")].copy()

def add_count(df):
    """Add 'Số lần gặp' per SĐT (non-empty)."""
    counts = df.groupby("SĐT", dropna=True).size().rename("Số lần gặp")
    df = df.merge(counts, on="SĐT", how="left")
    df["Số lần gặp"] = df["Số lần gặp"].fillna(0).astype(int)
    return df

def ensure_and_order_columns(df):
    """
    Keep ONLY these columns in this exact order (create empty ones if missing),
    and DROP all other columns.
    """
    desired = [
        "Ban", "Nhóm", "Mã TVV", "Tên TVV",
        "SĐT", "Tên KH", "Số lần gặp",
        "Ngày đăng ký", "Ngày thực hiện",
        "Ảnh báo cáo", "IP (trđ)", "ID"
    ]
    for c in desired:
        if c not in df.columns:
            df[c] = pd.NA
    return df[desired].copy()  # << return EXACTLY desired columns

def sort_preserving_sdt_groups(df):
    """
    - Non-empty SĐT rows grouped; within each SĐT group, sort by 'Ngày thực hiện' ascending.
    - Groups themselves ordered by 'Ban' ascending (group key = min Ban string).
    - Empty SĐT rows sorted by Ban then Ngày thực hiện and interleaved by Ban.
    """
    df = df.copy()
    df["_dt_th"] = pd.to_datetime(df["Ngày thực hiện"], errors="coerce")

    has_sdt = df["SĐT"].astype(str).str.strip().ne("") & df["SĐT"].notna()
    df_with = df.loc[has_sdt].copy()
    df_without = df.loc[~has_sdt].copy()

    groups = []
    for sdt_val, g in df_with.groupby("SĐT", sort=False):
        g_sorted = g.sort_values(by="_dt_th", na_position="last", kind="stable")
        ban_key = g_sorted["Ban"].astype(str).fillna("~").min()
        groups.append((ban_key, sdt_val, g_sorted))

    groups.sort(key=lambda t: (t[0], t[1]))  # by Ban asc, then SĐT

    df_without_sorted = df_without.sort_values(
        by=["Ban", "_dt_th"], na_position="last", kind="stable"
    )

    items = []
    for (ban_key, sdt_val, g_sorted) in groups:
        items.append((ban_key if ban_key is not None else "~", True, g_sorted))
    for _, row in df_without_sorted.iterrows():
        bkey = str(row["Ban"]) if pd.notna(row["Ban"]) else "~"
        items.append((bkey, False, row.to_frame().T))

    items.sort(key=lambda t: (t[0], not t[1]))  # keep groups before singletons on same Ban
    ordered = pd.concat([it[2] for it in items], ignore_index=True)
    if "_dt_th" in ordered.columns:
        ordered = ordered.drop(columns=["_dt_th"])
    return ordered

def merge_blocks_by_group(ws, group_col_idx, target_col_indices, start_row, end_row):
    """Merge vertically within each contiguous block sharing the same SĐT."""
    center = Alignment(horizontal="center", vertical="center")
    current_group = None
    block_start = None

    def merge_block(r1, r2):
        if r2 <= r1:
            return
        for c in target_col_indices:
            ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
            ws.cell(row=r1, column=c).alignment = center

    for r in range(start_row, end_row + 1):
        gval = ws.cell(row=r, column=group_col_idx).value
        if gval is None or str(gval).strip() == "":
            if current_group is not None and block_start is not None:
                merge_block(block_start, r - 1)
            current_group, block_start = None, None
            continue
        if gval != current_group:
            if current_group is not None and block_start is not None:
                merge_block(block_start, r - 1)
            current_group = gval
            block_start = r
    if current_group is not None and block_start is not None:
        merge_block(block_start, end_row)

# ---------- App flow ----------
if uploaded:
    xls = pd.ExcelFile(uploaded)
    st.write("Detected sheets:", xls.sheet_names)

    tvcn_sheet = st.selectbox(
        "TVCN sheet name",
        xls.sheet_names,
        index=(xls.sheet_names.index(tvcn_sheet_default)
               if tvcn_sheet_default in xls.sheet_names else 0),
    )
    kh_sheet = st.selectbox(
        "Khách hàng sheet name",
        xls.sheet_names,
        index=(xls.sheet_names.index(kh_sheet_default)
               if kh_sheet_default in xls.sheet_names else max(0, len(xls.sheet_names)-1)),
    )

    if st.button("Run merge"):
        try:
            df_tvcn = pd.read_excel(uploaded, sheet_name=tvcn_sheet, dtype=object, engine="openpyxl")
            df_kh   = pd.read_excel(uploaded, sheet_name=kh_sheet,   dtype=object, engine="openpyxl")

            # Validate
            missing = []
            if "ID" not in df_tvcn.columns:
                missing.append(f"{tvcn_sheet}.ID")
            for col in ["ID", "SĐT", "Tên KH"]:
                if col not in df_kh.columns:
                    missing.append(f"{kh_sheet}.{col}")
            if missing:
                st.error("Missing required columns: " + ", ".join(missing))
                st.stop()

            # Normalize & merge
            df_tvcn["ID_norm"] = df_tvcn["ID"].map(normalize_id)
            df_kh["ID_norm"]   = df_kh["ID"].map(normalize_id)
            df_kh["SĐT"]       = df_kh["SĐT"].map(normalize_phone)

            df_kh_subset = df_kh[["ID_norm", "SĐT", "Tên KH"]]
            df_merged = df_tvcn.merge(df_kh_subset, on="ID_norm", how="left").drop(columns=["ID_norm"])
            df_merged["SĐT"] = df_merged["SĐT"].astype("string")

            # Keep only Đã báo cáo
            df_filtered = filter_reported_rows(df_merged)

            # Count + select ONLY required columns in the specified order
            df_counted = add_count(df_filtered)
            df_ordered = ensure_and_order_columns(df_counted)

            # Sort (keep SĐT groups; per-group sort by Ngày thực hiện asc; groups by Ban asc)
            df_final = sort_preserving_sdt_groups(df_ordered)

            # Build workbook (replace TVCN with df_final)
            all_sheets = {name: pd.read_excel(uploaded, sheet_name=name, dtype=object, engine="openpyxl")
                          for name in xls.sheet_names}
            all_sheets[tvcn_sheet] = df_final

            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                for name, df in all_sheets.items():
                    df_to_write = df_final if name == tvcn_sheet else df
                    df_to_write.to_excel(writer, sheet_name=name, index=False)

                # Apply merges per SĐT group
                ws = writer.book[tvcn_sheet]
                headers = [cell.value for cell in ws[1]]
                sdt_col_idx = headers.index("SĐT") + 1
                slg_col_idx = headers.index("Số lần gặp") + 1

                data_start_row = 2
                data_end_row = 1 + len(df_final)

                merge_blocks_by_group(
                    ws,
                    group_col_idx=sdt_col_idx,
                    target_col_indices=[sdt_col_idx, slg_col_idx],
                    start_row=data_start_row,
                    end_row=data_end_row,
                )

            buf.seek(0)

            st.success(
                f"Done! Kept only 'Đã báo cáo' rows ({len(df_final)} rows). "
                "Kept ONLY the 12 required columns, ordered; SĐT as text; grouped, sorted, and merged per SĐT."
            )
            st.dataframe(df_final.head(100))
            st.download_button(
                label="Download updated Excel",
                data=buf,
                file_name="updated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            st.exception(e)
else:
    st.info("Please upload an .xlsx to begin.")
